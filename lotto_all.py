import collections
from openpyxl import load_workbook
wb = load_workbook("lotto_all.xlsx")
ws = wb.active

alpha = [ "J", "K", "L", "M", "N", "O"]
maxalpha = ["Q1", "R1", "S1", "T1", "U1", "V1"]
average1 = ["Q2", "R2", "S2", "T2", "U2", "V2"]
average2 = ["Q3", "R3", "S3", "T3", "U3", "V3"]
maxalpha2 = ["Q4", "R4", "S4", "T4", "U4", "V4"]
ocha1 = ["Q7", "Q8", "Q9", "Q10", "Q11", "Q12", "Q13", "Q14", "Q15", "Q16"]
ocha2 = ["R7", "R8", "R9", "R10", "R11", "R12", "R13", "R14", "R15", "R16"]
ocha3 = ["S7", "S8", "S9", "S10", "S11", "S12", "S13", "S14", "S15", "S16"]
ocha4 = ["T7", "T8", "T9", "T10", "T11", "T12", "T13", "T14", "T15", "T16"]
ocha5 = ["U7", "U8", "U9", "U10", "U11", "U12", "U13", "U14", "U15", "U16"]
ocha6 = ["V7", "V8", "V9", "V10", "V11", "V12", "V13", "V14", "V15", "V16"]
ocha = [ocha1, ocha2, ocha3, ocha4, ocha5, ocha6]
s = ["W1", "W2", "W3"]

while True:
    for i in range(6):
        전체숫자 = []
        자리전체숫자 = []
        최근숫자 = []
        다음숫자 = []
        부호 = []
        for j in range(6):
            for row in ws.rows:
                전체숫자.append(row[j+1].value)

        for row in ws.rows:
            자리전체숫자.append(row[i+1].value)

        최근숫자.append(자리전체숫자[0])
        if 자리전체숫자[0] == 자리전체숫자[1]:
            다음숫자.append(자리전체숫자[0])
        자리전체숫자[0] = 0

        while 최근숫자[0] in 전체숫자:
            a = 전체숫자.index(최근숫자[0])
            if a != 0:
                다음숫자.append(전체숫자[a-1])
            if a + 1 < len(전체숫자):
                if 전체숫자[a+1] == 전체숫자[a]:
                    다음숫자.append(전체숫자[a])
                    전체숫자[a+1] = 0
            전체숫자[a] = 0

        while 0 in 다음숫자:
            다음숫자.remove(0)
        

        for j in range(len(다음숫자)):
            ws[alpha[i]][j].value = 다음숫자[j]

        for q in range(1,11):
            ws[ocha[i][q-1]] = 다음숫자[q-1] - 다음숫자[q]

        counts = collections.Counter(다음숫자)

        ws[maxalpha[i]].value = counts.most_common(1)[0][0]
        ws[average1[i]].value = "=AVERAGE({}1:{}{})".format(alpha[i], alpha[i],len(다음숫자)-1)
        ws[average2[i]].value = "=AVERAGE({},{})".format(maxalpha[i], average1[i])

    break
for i in range(1,4):
    ws[s[i-1]] = "=SUM(Q{}:V{})".format(i,i)
j = 1
while True:
    ws["P{}".format(j)] = "=SUM(J{}:O{})".format(j, j)
    j += 1
    if ws["J{}".format(j)].value == None:
        break

wb.save("lotto_all.xlsx")
wb.close()