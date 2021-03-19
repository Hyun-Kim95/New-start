import collections
from openpyxl import load_workbook
wb = load_workbook("lotto.xlsx")
ws = wb.active

alpha = [ "J", "K", "L", "M", "N", "O"]
maxalpha = ["Q1", "R1", "S1", "T1", "U1", "V1"]
average1 = ["Q2", "R2", "S2", "T2", "U2", "V2"]
average2 = ["Q3", "R3", "S3", "T3", "U3", "V3"]
maxalpha2 = ["Q4", "R4", "S4", "T4", "U4", "V4"]
s = ["W1", "W2", "W3"]

while True:
    for i in range(6):
        전체숫자 = []
        최근숫자 = []
        다음숫자 = []
        부호 = []
        for row in ws.rows:
            전체숫자.append(row[i+1].value)

        최근숫자.append(전체숫자[0])
        if 전체숫자[0] == 전체숫자[1]:
            다음숫자.append(전체숫자[0])
        전체숫자[0] = 0
        while 최근숫자[0] in 전체숫자:
            a = 전체숫자.index(최근숫자[0])
            다음숫자.append(전체숫자[a-1])
            if a + 1 < len(전체숫자):
                if 전체숫자[a+1] == 전체숫자[a]:
                    다음숫자.append(전체숫자[a])
                    전체숫자[a+1] = 0
            전체숫자[a] = 0

        while 0 in 다음숫자:
            다음숫자.remove(0)
        

        for j in range(len(다음숫자)):
            ws[alpha[i]][j].value = 다음숫자[j]

        counts = collections.Counter(다음숫자)

        ws[maxalpha[i]].value = counts.most_common(1)[0][0]
        ws[average1[i]].value = "=AVERAGE({}1:{}{})".format(alpha[i], alpha[i],len(다음숫자)-1)
        ws[average2[i]].value = "=AVERAGE({},{})".format(maxalpha[i], average1[i])

    break
for i in range(1,4):
    ws[s[i-1]] = "=SUM(Q{}:V{})".format(i,i)
j = 1
while True:
    ws["P{}".format(j)] = "=SUM(J{}:O{})".format(j, j)
    j += 1
    if ws["J{}".format(j)].value == None:
        break

wb.save("lotto1.xlsx")
wb.close()